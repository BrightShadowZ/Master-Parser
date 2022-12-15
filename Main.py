# Import the necessary libraries
import zipfile

import openpyxl
import os

# Open the Master sheet
wb = openpyxl.load_workbook("Master sheet.xlsx")
sheet = wb.active
folder_path = "\\Folder Path\\"

# Read the data into a dictionary, with the PRODUCT CODE as the key
data = {}

for row in sheet.iter_rows(min_row=0, max_col=7, values_only=True):
    customer_code = row[0]
    product_code = row[1]
    data[customer_code, product_code] = row[0:]
num = 17
lines = 0

# Loop through each Excel file in the folder
for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx'):
        try:
            wb = openpyxl.load_workbook(folder_path + filename)
            sheet = wb.active
            sheet.delete_cols(20, 6)
            sheet['T16'] = 'Std cost'
            sheet['U16'] = 'Quantity Sold'
            sheet['V16'] = 'ring Trading Margin'
            sheet['W16'] = 'Osram Net Sales'
            sheet['X16'] = 'Category (ex D1A-halogen)'

            for row in sheet.iter_rows(min_row=17, max_col=0, values_only=True):

                product_code = row[0], row[1]

                try:
                    if product_code in data:
                        # Copy the data from the Master sheet into the appropriate cells

                        sheet['T' + str(num)] = data[product_code][6]
                        sheet['U' + str(num)] = data[product_code][5]
                        sheet['V' + str(num)] = data[product_code][4]
                        sheet['W' + str(num)] = data[product_code][3]
                        sheet['X' + str(num)] = data[product_code][2]
                        num += 1
                        lines += 1
                    else:
                        num += 1
                        lines += 1
                        continue

                except Exception as e:
                    print(f"Error updating file {filename}: {e}")
                    continue
            # Save the Excel file
            try:
                wb.save(folder_path + filename)
                num -= lines
                lines -= lines

                print(f"Updated file {filename}")
            except Exception as e:
                print(f"Error saving file {filename}: {e}")
        except zipfile.BadZipFile:
            # Handle the error if the file is not a valid Excel file
            print(f"{filename} is not a valid Excel file, skipping...")
